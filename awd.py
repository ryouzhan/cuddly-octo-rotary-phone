# -*- coding: utf-8 -*-
"""
AWD 发货单智能生成（从原 AWDSmartGeneratorApp.start_process / beautify_sheet 剥离 UI 后保留的业务逻辑）
"""
import os
import traceback
from datetime import datetime
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side


def get_region(addr):
    """从配送地址识别仓库区域（AWD 用法）"""
    addr = str(addr).upper()
    if 'AWD' in addr:
        return "AWD仓"
    east = ['ABE8', 'AVP1', 'DCA6', 'TEB9', 'PHL7', 'BDL3']
    central = ['DFW6', 'FOE1', 'MDW2', 'IND7', 'MEM1']
    west = ['IUTE', 'ONT8', 'LAS1', 'LAX9', 'GYR2', 'ABQ2', 'SCK4', 'SMF3']
    if any(x in addr for x in east):
        return "美东"
    if any(x in addr for x in central):
        return "美中"
    if any(x in addr for x in west):
        return "美西"
    return "其他"


def beautify_sheet(ws):
    """AWD 表格美化：表头深蓝白字 + 隔行浅灰底色 + 自动列宽"""
    header_fill = PatternFill(start_color="203764", end_color="203764", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    stripe_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin', color='BFBFBF'),
        right=Side(style='thin', color='BFBFBF'),
        top=Side(style='thin', color='BFBFBF'),
        bottom=Side(style='thin', color='BFBFBF')
    )
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row)):
        for cell in row:
            cell.border = thin_border
            cell.alignment = center_align
            if i == 0:
                cell.fill = header_fill
                cell.font = header_font
            elif i % 2 == 0:
                cell.fill = stripe_fill
    for col in ws.columns:
        max_len = 0
        column = col[0].column_letter
        for cell in col:
            try:
                val_len = len(str(cell.value).encode('gbk'))
                if val_len > max_len:
                    max_len = val_len
            except Exception:
                pass
        ws.column_dimensions[column].width = min(max_len + 4, 50)


def process(awd_path, comm_path, output_path, log=None, progress=None):
    """
    AWD 主处理流程
    :param awd_path:   备货单表格
    :param comm_path:  商品列表
    :param output_path:处理结果保存路径
    :param log:        日志回调
    :param progress:   进度回调 0-100
    :return: output_path
    """
    def emit(msg):
        (log or print)(msg)

    def prog(p, msg=""):
        if progress:
            progress(p, msg)
        if msg:
            emit(msg)

    try:
        prog(5, "读取并清洗数据（清除隐形空格）...")

        ext_awd = os.path.splitext(awd_path)[1].lower()
        df_awd = pd.read_csv(awd_path) if ext_awd == '.csv' else pd.read_excel(awd_path)
        df_awd.rename(columns=lambda x: str(x).strip(), inplace=True)

        target_cols = ['商品品牌', '箱规长(cm)', '箱规宽(cm)', '箱规高(cm)', '单箱重量(kg)', '单箱数量(pcs)', '供应商名称', '采购单价', '采购成本(￥)']

        ext_comm = os.path.splitext(comm_path)[1].lower()
        if ext_comm == '.csv':
            df_comm = pd.read_csv(comm_path)
            df_comm.rename(columns=lambda x: str(x).strip(), inplace=True)
        else:
            xls = pd.ExcelFile(comm_path)
            best_sheet = xls.sheet_names[0]
            max_match = 0
            for sheet in xls.sheet_names:
                df_tmp = pd.read_excel(xls, sheet_name=sheet, nrows=0)
                df_tmp.rename(columns=lambda x: str(x).strip(), inplace=True)
                match = sum(1 for c in target_cols if c in df_tmp.columns)
                if match > max_match:
                    max_match = match
                    best_sheet = sheet
            emit(f"自动锁定 Commodities 数据源工作表：[{best_sheet}]")
            df_comm = pd.read_excel(xls, sheet_name=best_sheet)
            df_comm.rename(columns=lambda x: str(x).strip(), inplace=True)

        if 'SKU' not in df_awd.columns or 'SKU' not in df_comm.columns:
            raise ValueError("找不到 'SKU' 列，请检查表格是否标准！")

        prog(30, "启动 SKU 智能映射引擎 (穿透 ZF- 前缀)...")
        df_awd['Merge_SKU'] = df_awd['SKU'].astype(str).str.strip()
        df_comm['Merge_SKU'] = df_comm['SKU'].astype(str).str.strip()
        df_comm = df_comm.drop_duplicates(subset=['Merge_SKU'])

        existing_cols = [c for c in target_cols if c in df_comm.columns]
        comm_dict = df_comm.set_index('Merge_SKU')[existing_cols].to_dict('index')

        merged_rows = []
        match_count = 0
        for _, row in df_awd.iterrows():
            awd_sku = str(row['Merge_SKU'])
            no_zf_sku = awd_sku[3:] if awd_sku.upper().startswith('ZF-') else awd_sku
            match_data = {}
            if awd_sku in comm_dict:
                match_data = comm_dict[awd_sku]
                match_count += 1
            elif no_zf_sku in comm_dict:
                match_data = comm_dict[no_zf_sku]
                match_count += 1
            combined = {**row.to_dict(), **match_data}
            merged_rows.append(combined)
        df_merged = pd.DataFrame(merged_rows)
        emit(f"匹配完成！共成功匹配 {match_count} / {len(df_awd)} 条商品数据。")

        prog(55, "正在执行逐行智能数据补全...")
        df_detail = pd.DataFrame()
        df_detail['SKU'] = df_merged.get('SKU', '')
        df_detail['品名'] = df_merged.get('品名', '')
        df_detail['商品图片'] = df_merged.get('图片', '')

        # 发货量
        f_qty = pd.Series(0, index=df_merged.index)
        if '申报量' in df_merged.columns:
            f_qty = pd.to_numeric(df_merged['申报量'], errors='coerce').fillna(0)
        if '备货量' in df_merged.columns:
            b_qty = pd.to_numeric(df_merged['备货量'], errors='coerce').fillna(0)
            f_qty = np.where(f_qty > 0, f_qty, b_qty)
        df_detail['发货量'] = f_qty

        # 单箱数量
        p_qty = pd.Series(0, index=df_merged.index)
        if '单箱数量(pcs)' in df_merged.columns:
            p_qty = pd.to_numeric(df_merged['单箱数量(pcs)'], errors='coerce').fillna(0)
        if '单箱数量' in df_merged.columns:
            fallback_p = pd.to_numeric(df_merged['单箱数量'], errors='coerce').fillna(0)
            p_qty = np.where(p_qty > 0, p_qty, fallback_p)
        df_detail['单箱数量'] = p_qty

        # 箱数
        safe_p_qty = np.where(df_detail['单箱数量'] > 0, df_detail['单箱数量'], 1)
        df_detail['箱数'] = np.where(df_detail['单箱数量'] > 0, np.ceil(df_detail['发货量'] / safe_p_qty), 0)

        # 配送地址
        addr_code = pd.Series('', index=df_merged.index)
        for col in ['物流中心编码', '收货仓库']:
            if col in df_merged.columns:
                a_col = df_merged[col].fillna('')
                addr_code = np.where(addr_code != '', addr_code, a_col)
        df_detail['配送地址'] = addr_code

        # 货件编号
        h_code = pd.Series('', index=df_merged.index)
        if '货件编号' in df_merged.columns:
            h_code = df_merged['货件编号'].fillna('')
        if '单据备注' in df_merged.columns:
            extracted = df_merged['单据备注'].astype(str).str.extract(r'货件编号[：:]([A-Za-z0-9\-]+)')[0].fillna('')
            h_code = np.where(h_code != '', h_code, extracted)
        if '备货单号' in df_merged.columns:
            b_code = df_merged['备货单号'].fillna('')
            h_code = np.where(h_code != '', h_code, b_code)
        df_detail['货件编号'] = h_code

        df_detail['供应商'] = df_merged.get('供应商名称', '')
        df_detail['ReferenceId'] = df_merged.get('ReferenceId', '')
        df_detail['箱号'] = ''

        # 总箱数编号
        z_code = pd.Series('', index=df_merged.index)
        if '申报量(货件)' in df_merged.columns:
            z_code = df_merged['申报量(货件)'].fillna('')
        if '备货单号' in df_merged.columns:
            b_code2 = df_merged['备货单号'].fillna('')
            z_code = np.where(z_code != '', z_code, b_code2)
        df_detail['总箱数编号'] = z_code

        # 箱规 / 体积
        w = pd.to_numeric(df_merged.get('单箱重量(kg)', 0), errors='coerce').fillna(0)
        l = pd.to_numeric(df_merged.get('箱规长(cm)', 0), errors='coerce').fillna(0)
        wd = pd.to_numeric(df_merged.get('箱规宽(cm)', 0), errors='coerce').fillna(0)
        h = pd.to_numeric(df_merged.get('箱规高(cm)', 0), errors='coerce').fillna(0)

        df_detail['外箱重量(kg)'] = w.round(2)
        df_detail['外箱总重量(kg)'] = (w * df_detail['箱数']).round(2)
        df_detail['外箱长(cm)'] = l.round(2)
        df_detail['外箱宽(cm)'] = wd.round(2)
        df_detail['外箱高(cm)'] = h.round(2)
        vol = (l * wd * h) / 1000000
        df_detail['外箱体积(m³)'] = vol.round(2)
        df_detail['外箱总体积(m³)'] = (vol * df_detail['箱数']).round(2)
        df_detail['外箱总体积重(kg)'] = (df_detail['外箱总体积(m³)'] * 167).round(2)

        # 时间
        df_detail['创建时间'] = df_merged.get('创建时间', '')
        time_code = pd.Series('', index=df_merged.index)
        for col in ['预计发货日期', '实际发货时间', '预计到货时间']:
            if col in df_merged.columns:
                t_col = df_merged[col].fillna('')
                time_code = np.where(time_code != '', time_code, t_col)
        df_detail['发货时间'] = time_code
        df_detail['物流商'] = ''

        # 货值核算
        u_price = pd.Series(0.0, index=df_merged.index)
        price_columns = ['采购单价(CNY)', '指定采购单价(CNY)', '采购成本(￥)', '采购单价', '单价']
        for col in price_columns:
            if col in df_merged.columns:
                temp_p = pd.to_numeric(df_merged[col], errors='coerce').fillna(0)
                u_price = np.where(u_price > 0, u_price, temp_p)
        df_detail['单价'] = u_price.round(2)
        df_detail['总货值'] = (df_detail['发货量'] * df_detail['单价']).round(2)

        prog(75, "正在按货件编号生成汇总表...")
        df_detail['仓库分区'] = df_detail['配送地址'].apply(get_region)
        summary_group = df_detail.groupby('货件编号').agg({
            '仓库分区': 'first',
            '配送地址': 'first',
            'ReferenceId': 'first',
            '品名': 'first',
            '供应商': 'first',
            '箱数': 'sum',
            '外箱总重量(kg)': 'sum',
            '外箱总体积(m³)': 'sum',
            '外箱总体积重(kg)': 'sum',
            '总货值': 'sum'
        }).reset_index()
        df_summary = summary_group[['仓库分区', '配送地址', '货件编号', 'ReferenceId', '品名', '供应商', '箱数', '外箱总重量(kg)', '外箱总体积(m³)', '外箱总体积重(kg)', '总货值']]
        df_summary.columns = ['仓库分区', '配送地址', '货件编号', 'ReferenceId', '品名', '供应商', '总箱数', '外箱总重量', '外箱总体积', '外箱总体积重(kg)', '货件总货值']

        prog(90, "正在保存并美化表格...")
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df_summary.to_excel(writer, sheet_name='汇总结果', index=False)
            df_detail.drop(columns=['仓库分区']).to_excel(writer, sheet_name='详细数据', index=False)
        wb = load_workbook(output_path)
        beautify_sheet(wb['汇总结果'])
        beautify_sheet(wb['详细数据'])
        wb.save(output_path)

        prog(100, f"✅ 货值计算完毕！文件已保存：{output_path}")
        return output_path
    except Exception as e:
        emit(f"【报错】: {str(e)}")
        emit(traceback.format_exc())
        raise
