# -*- coding: utf-8 -*-
"""
FBA 仓库精准映射与表格美化（从原程序 1-411 行剥离 UI 后保留的业务函数）
"""
import re
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

# ===================== 379 个 FBA 仓库精准映射表 =====================
FBA_WAREHOUSE_MAP = {
    # 美西仓库
    "PHX3": ("AZ", "美西"), "PHX5": ("AZ", "美西"), "PHX6": ("AZ", "美西"), "PHX7": ("AZ", "美西"),
    "ONT2": ("CA", "美西"), "ONT6": ("CA", "美西"), "ONT8": ("CA", "美西"), "ONT9": ("CA", "美西"),
    "ONT7": ("CA", "美西"), "OAK3": ("CA", "美西"), "OAK4": ("CA", "美西"), "LGB3": ("CA", "美西"),
    "LGB4": ("CA", "美西"), "LGB6": ("CA", "美西"), "LGB7": ("CA", "美西"), "LGB8": ("CA", "美西"),
    "LGB9": ("CA", "美西"), "LGB1": ("CA", "美西"), "SNA4": ("CA", "美西"), "SJC7": ("CA", "美西"),
    "SJC8": ("CA", "美西"), "SMF1": ("CA", "美西"), "SMF3": ("CA", "美西"), "SMF6": ("CA", "美西"),
    "SMF7": ("CA", "美西"), "SLC1": ("UT", "美西"), "SLC2": ("UT", "美西"), "SLC3": ("UT", "美西"),
    "RNO1": ("NV", "美西"), "RNO2": ("NV", "美西"), "RNO4": ("NV", "美西"), "LAS1": ("NV", "美西"),
    "LAS2": ("NV", "美西"), "LAS6": ("NV", "美西"), "LAS7": ("NV", "美西"), "BFI1": ("WA", "美西"),
    "BFI3": ("WA", "美西"), "BFI4": ("WA", "美西"), "BFI5": ("WA", "美西"), "BFI7": ("WA", "美西"),
    "BFI9": ("WA", "美西"), "SEA6": ("WA", "美西"), "SEA8": ("WA", "美西"), "OLM1": ("WA", "美西"),
    "GEG1": ("WA", "美西"), "GEG2": ("WA", "美西"), "PSC2": ("WA", "美西"), "PDX6": ("OR", "美西"),
    "PDX7": ("OR", "美西"), "PDX9": ("OR", "美西"), "FAT1": ("CA", "美西"), "FAT2": ("CA", "美西"),
    "SBD1": ("CA", "美西"), "SBD2": ("CA", "美西"), "SBD3": ("CA", "美西"), "SCK1": ("CA", "美西"),
    "SCK3": ("CA", "美西"), "SCK4": ("CA", "美西"), "SCK8": ("CA", "美西"), "KRB1": ("CA", "美西"),
    "KRB4": ("CA", "美西"), "KRB7": ("CA", "美西"), "XIX6": ("CA", "美西"), "XIX7": ("CA", "美西"),
    "XLX2": ("CA", "美西"), "XLX1": ("CA", "美西"), "XLX3": ("CA", "美西"), "QXY9": ("CA", "美西"),
    "QXY5": ("CA", "美西"), "MCE1": ("CA", "美西"), "MIT2": ("CA", "美西"), "MIT2-UPS": ("CA", "美西"),
    "BFL1": ("CA", "美西"), "BFL2": ("CA", "美西"), "ABS4": ("CA", "美西"), "JOT1": ("CA", "美西"),
    "IUSQ": ("CA", "美西"), "IUTI": ("CA", "美西"), "XLG1": ("CA", "美西"), "POC1": ("CA", "美西"),
    "POC2": ("CA", "美西"), "POC3": ("CA", "美西"), "POC3-UPS": ("CA", "美西"), "HLI2": ("CA", "美西"),
    "TCY1": ("CA", "美西"), "TCY2": ("CA", "美西"), "AKC1": ("OH", "美中"), "AZA4": ("AZ", "美西"),
    "GEU3": ("AZ", "美西"), "GEU5": ("AZ", "美西"), "GEU2": ("AZ", "美西"), "GYR1": ("AZ", "美西"),
    "GYR2": ("AZ", "美西"), "GYR3": ("AZ", "美西"), "PCA1": ("AZ", "美西"), "TUS2": ("AZ", "美西"),
    "ABQ2": ("NM", "美西"), "ABQ2-UPS": ("AZ", "美西"), "BOI2": ("ID", "美西"), "VGT2": ("NV", "美西"),
    "VGT2-UPS": ("CA", "美西"), "IUSP": ("CA", "美西"), "IUSF": ("TX", "美中"), "IUSL": ("MD", "美东"),
    "IUSJ": ("CA", "美西"), "IUST": ("PA", "美东"), "IUSR": ("SC", "美东"), "IUTH": ("TX", "美中"),
    # 美中仓库
    "ICT2": ("KS", "美中"), "IND1": ("IN", "美中"), "IND2": ("IN", "美中"), "IND3": ("IN", "美中"),
    "IND4": ("IN", "美中"), "IND5": ("IN", "美中"), "IND6": ("IN", "美中"), "IND9": ("IN", "美中"),
    "SDF8": ("IN", "美中"), "TUL1": ("OK", "美中"), "TUL2": ("OK", "美中"), "MCI1": ("KS", "美中"),
    "MKC4": ("KS", "美中"), "MKC6": ("KS", "美中"), "MCI3": ("MO", "美中"), "FOE1": ("KS", "美中"),
    "LEX1": ("KY", "美中"), "LEX2": ("KY", "美中"), "SDF1": ("KY", "美中"), "SDF2": ("NY", "美东"),
    "SDF4": ("KY", "美中"), "SDF6": ("KY", "美中"), "SDF7": ("KY", "美中"), "SDF9": ("KY", "美中"),
    "CVG1": ("KY", "美中"), "CVG2": ("KY", "美中"), "CVG3": ("KY", "美中"), "CVG5": ("KY", "美中"),
    "IVSA": ("KY", "美中"), "IVSB": ("KY", "美中"), "CMH1": ("OH", "美中"), "CMH1-11999": ("OH", "美中"),
    "CMH2": ("OH", "美中"), "CMH3": ("OH", "美中"), "CMH4": ("OH", "美中"), "CMH6": ("OH", "美中"),
    "CMH7": ("OH", "美中"), "CLE2": ("OH", "美中"), "CLE3": ("OH", "美中"), "AKR1": ("OH", "美中"),
    "DFW6": ("TX", "美中"), "DFW7": ("TX", "美中"), "DFW8": ("TX", "美中"), "DS-DFW6": ("CA", "美西"),
    "IAH3": ("TX", "美中"), "HOU1": ("TX", "美中"), "HOU2": ("TX", "美中"), "HOU3": ("TX", "美中"),
    "HOU7": ("TX", "美中"), "HOU8": ("TX", "美中"), "XUSB": ("TX", "美中"), "FTW1": ("TX", "美中"),
    "FTW2": ("TX", "美中"), "FTW3": ("TX", "美中"), "FTW5": ("TX", "美中"), "FTW6": ("TX", "美中"),
    "FTW9": ("TX", "美中"), "SAT1": ("TX", "美中"), "SAT2": ("TX", "美中"), "SAT4": ("TX", "美中"),
    "SAT6": ("TX", "美中"), "AUS1": ("TX", "美中"), "AUS2": ("TX", "美中"), "DAL2": ("TX", "美中"),
    "DAL3": ("TX", "美中"), "AMA1": ("TX", "美中"), "AFW1": ("TX", "美中"), "KTX2": ("TX", "美中"),
    "ITX1": ("TX", "美中"), "ITX2": ("TX", "美中"), "OKC1": ("OK", "美中"), "OKC2": ("OK", "美中"),
    "ELP1": ("TX", "美中"), "ELP4": ("TX", "美中"), "LIT2": ("AR", "美中"), "MKE1": ("WI", "美中"),
    "MKE2": ("WI", "美中"), "JVL1": ("WI", "美中"), "FWA4": ("IN", "美中"), "MQJ1": ("IN", "美中"),
    "MQJ2": ("IN", "美中"), "PPO4": ("IN", "美中"), "PPO4(46410)": ("IN", "美中"), "PPO4-UPS": ("IN", "美中"),
    "MDW2": ("IL", "美中"), "MDW2-250": ("IL", "美中"), "MDW2-402": ("IL", "美中"), "MDW4": ("IL", "美中"),
    "MDW4-60433": ("IL", "美中"), "MDW5": ("IL", "美中"), "MDW6": ("IL", "美中"), "MDW7": ("IL", "美中"),
    "MDW8": ("IL", "美中"), "MDW9": ("IL", "美中"), "ORD2": ("IL", "美中"), "ORD6": ("IL", "美中"),
    "STL4": ("IL", "美中"), "STL3": ("MO", "美中"), "STL6": ("IL", "美中"), "STL8": ("MO", "美中"),
    "DSM5": ("IA", "美中"), "MSP1": ("MN", "美中"), "MSP6": ("MN", "美中"), "DTW1": ("MI", "美中"),
    "DTW3": ("MI", "美中"), "DET1": ("MI", "美中"), "DET2": ("MI", "美中"), "GRR1": ("MI", "美中"),
    "SMI1": ("MI", "美中"), "LAN2": ("MI", "美中"), "LAN2-UPS": ("MI", "美中"), "FAR1": ("ND", "美中"),
    "DEN2": ("CO", "美中"), "DEN3": ("CO", "美中"), "DEN7": ("CO", "美中"), "DEN8": ("CO", "美中"),
    "PCW1": ("OH", "美中"), "RFD2": ("IL", "美中"), "RFD4": ("IL", "美中"),
    # 美东仓库
    "ABE2": ("PA", "美东"), "ABE3": ("PA", "美东"), "ABE4": ("PA", "美东"), "ABE5": ("PA", "美东"),
    "ABE8": ("NJ", "美东"), "ABE8-309": ("NJ", "美东"), "AVP1": ("PA", "美东"), "AVP3": ("PA", "美东"),
    "AVP9": ("PA", "美东"), "PHL1": ("DE", "美东"), "PHL3": ("DE", "美东"), "PHL4": ("PA", "美东"),
    "PHL5": ("PA", "美东"), "PHL6": ("PA", "美东"), "PHL7": ("DE", "美东"), "PHL8": ("DE", "美东"),
    "EWR4": ("NJ", "美东"), "EWR5": ("NJ", "美东"), "EWR7": ("NJ", "美东"), "EWR9": ("NJ", "美东"),
    "EWR2": ("NJ", "美东"), "TEB3": ("NJ", "美东"), "TEB6": ("NJ", "美东"), "TEB4": ("NJ", "美东"),
    "TEB9": ("NJ", "美东"), "ACY1": ("NJ", "美东"), "ACY2": ("NJ", "美东"), "EL96": ("NJ", "美东"),
    "TTN2": ("NJ", "美东"), "XEW2": ("GA", "美东"), "XEW5": ("NJ", "美东"), "JFK8": ("NY", "美东"),
    "LGA9": ("NJ", "美东"), "ALB1": ("NY", "美东"), "SWF1": ("NY", "美东"), "SWF2": ("NY", "美东"),
    "SWF2-UPS": ("NY", "美东"), "GGE1": ("MA", "美东"), "GGE2": ("NY", "美东"), "GGE3": ("CT", "美东"),
    "GGE4": ("NY", "美东"), "GGE5": ("NY", "美东"), "GGE6": ("NY", "美东"), "GGE7": ("NY", "美东"),
    "GGE8": ("VA", "美东"), "BOS1": ("NH", "美东"), "BOS2": ("MA", "美东"), "BOS7": ("MA", "美东"),
    "BDL1": ("CT", "美东"), "BDL2": ("CT", "美东"), "BDL3": ("CT", "美东"), "BDL6": ("CT", "美东"),
    "PVD1": ("RI", "美东"), "MHT1": ("NH", "美东"), "TPA1": ("FL", "美东"), "TPA2": ("FL", "美东"),
    "TPA3": ("FL", "美东"), "TPA6": ("FL", "美东"), "MIA1": ("FL", "美东"), "MIA1-33054": ("FL", "美东"),
    "JAX3": ("FL", "美东"), "JAX2": ("FL", "美东"), "JAX7": ("FL", "美东"), "MCO1": ("FL", "美东"),
    "MCO2": ("FL", "美东"), "RSW1": ("FL", "美东"), "RSW2": ("FL", "美东"), "RSW3": ("FL", "美东"),
    "PBI1": ("FL", "美东"), "PBI2": ("FL", "美东"), "PBI3": ("FL", "美东"), "PBI3-UPS": ("FL", "美东"),
    "TMB8": ("FL", "美东"), "TMB8-UPS": ("FL", "美东"), "WBW2": ("PA", "美东"), "WBW2-UPS": ("PA", "美东"),
    "HIA1": ("PA", "美东"), "HGR2": ("MD", "美东"), "HGR6": ("MD", "美东"), "XLX7": ("CA", "美西"),
    "RIC1": ("VA", "美东"), "RIC2": ("VA", "美东"), "RIC3": ("VA", "美东"), "RYY2": ("GA", "美东"),
    "RYY2-UPS": ("GA", "美东"), "ORF2": ("VA", "美东"), "ORF2-UPS": ("PA", "美东"), "BWI1": ("VA", "美东"),
    "BWI2": ("MD", "美东"), "BWI4": ("VA", "美东"), "DCA1": ("MD", "美东"), "DCA6": ("MD", "美东"),
    "CLT2": ("NC", "美东"), "CLT3": ("NC", "美东"), "CLT4": ("NC", "美东"), "CLT6": ("NC", "美东"),
    "RDU1": ("NC", "美东"), "RDU2": ("NC", "美东"), "RDU4": ("NC", "美东"), "GSO1": ("NC", "美东"),
    "CAE1": ("SC", "美东"), "GSP1": ("SC", "美东"), "MYR1": ("SC", "美东"), "MYR2": ("SC", "美东"),
    "SAV1": ("SC", "美东"), "SAV2": ("SC", "美东"), "SAV3": ("GA", "美东"), "ATL6": ("GA", "美东"),
    "ATL2": ("GA", "美东"), "ATL7": ("GA", "美东"), "ATL8": ("GA", "美东"), "MGE3": ("GA", "美东"),
    "MGE1": ("GA", "美东"), "CSG1": ("GA", "美东"), "AGS1": ("GA", "美东"), "XAV3": ("GA", "美东"),
    "CHA1": ("TN", "美中"), "CHA2": ("TN", "美中"), "BNA1": ("TN", "美中"), "BNA2": ("TN", "美中"),
    "BNA3": ("TN", "美中"), "BNA4": ("TN", "美中"), "BNA6": ("TN", "美中"), "BNA6-UPS": ("TN", "美中"),
    "MEM1": ("TN", "美中"), "MEM2": ("MS", "美东"), "MEM4": ("TN", "美中"), "MEM6": ("MS", "美东"),
    "MEM8": ("MS", "美东"), "LFT1": ("LA", "美东"), "HSP1": ("AL", "美东"), "HSV2": ("AL", "美东"),
    "HSP0": ("AL", "美东"), "BHM1": ("AL", "美东"), "LBE1": ("PA", "美东"), "LBE1-UPS": ("PA", "美东"),
    "RDG1": ("PA", "美东"), "MID1": ("PA", "美东"), "HEA2": ("PA", "美东"), "CHO1": ("VA", "美东"),
    "RMN3": ("VA", "美东"), "RMN3-UPS": ("PA", "美东"), "PSC2-UPS": ("OR", "美西"), "KRB5": ("MD", "美东"),
    "KRB2": ("VA", "美东"), "XRI3": ("VA", "美东"), "XLX6": ("NC", "美东"), "LIB1": ("PA", "美东")
}


def beautify_excel(file_path, col_width=15, row_height=20, log=None):
    """美化 Excel 表格，按货件编号设置行底色"""
    def emit(msg):
        (log or print)(msg)
    try:
        wb = load_workbook(file_path)
        font = Font(name='Calibri', size=11, bold=True)
        alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin = Side(border_style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        colors = [
            "FFCCCC", "CCFFCC", "CCCCFF", "FFFFCC", "CCFFFF", "FFCCFF",
            "FF9999", "99FF99", "9999FF", "FFCC99", "99FFCC", "CCCC99"
        ]
        for sheet_name in ['详细数据', '汇总结果']:
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            for col in ws.columns:
                column_letter = col[0].column_letter
                ws.column_dimensions[column_letter].width = col_width
            for row in ws.iter_rows():
                if not row[0].row:
                    continue
                ws.row_dimensions[row[0].row].height = row_height
                for cell in row:
                    if cell.value is None:
                        continue
                    cell.alignment = alignment
                    cell.border = border
                    if cell.row == 1:
                        cell.font = font
            if sheet_name == '详细数据':
                header_row = [cell.value for cell in ws[1]]
                if "货件编号" not in header_row:
                    continue
                item_code_col_idx = header_row.index("货件编号") + 1
                item_codes = {}
                for row in ws.iter_rows(min_row=2, min_col=item_code_col_idx, max_col=item_code_col_idx):
                    val = row[0].value
                    if val and val not in item_codes:
                        item_codes[val] = None
                unique_codes = list(item_codes.keys())
                for idx, code in enumerate(unique_codes):
                    item_codes[code] = colors[idx % len(colors)]
                for row in ws.iter_rows(min_row=2):
                    code_cell = row[item_code_col_idx - 1]
                    code_val = code_cell.value
                    if code_val in item_codes:
                        fill = PatternFill(
                            start_color=item_codes[code_val],
                            end_color=item_codes[code_val],
                            fill_type="solid"
                        )
                        for cell in row:
                            cell.fill = fill
        wb.save(file_path)
        emit(f"表格美化完成: {file_path}")
    except Exception as e:
        emit(f"美化出错: {str(e)}")


def load_and_merge_data(file_path, log=None):
    """加载发货单详情 + 装箱信息并合并"""
    def emit(msg):
        (log or print)(msg)
    try:
        xl_file = pd.ExcelFile(file_path)
        sheet_names = xl_file.sheet_names
        emit(f"读取到工作表: {sheet_names}")
        required_sheets = ['发货单详情', '装箱信息']
        for sheet in required_sheets:
            if sheet not in sheet_names:
                raise KeyError(f"文件缺少必须工作表：{sheet}")
        df_detail = pd.read_excel(file_path, sheet_name='发货单详情')
        df_pack = pd.read_excel(file_path, sheet_name='装箱信息')
        df_detail.rename(columns={
            '货件编号(发货商品)': '货件编号',
            '发货单号(基本信息)': '发货单号',
            'SKU(发货商品)': 'SKU',
            '发货量(发货商品)': '发货量',
            '品名(发货商品)': '品名',
            '商品图片(发货商品)': '商品图片',
            '商品品牌': '供应商'
        }, inplace=True)
        df_pack.rename(columns={
            '货件编号(装箱信息)': '货件编号',
            '发货单号(装箱信息)': '发货单号',
            'SKU(装箱信息)': 'SKU',
            '发货量(装箱信息)': '发货量',
            'ReferenceId(装箱信息)': 'ReferenceId'
        }, inplace=True)
        merged = pd.merge(
            df_detail, df_pack,
            on=['货件编号', '发货单号', 'SKU', '发货量'],
            how='left'
        )
        keep_cols = [
            'SKU', '品名', '商品图片', '发货量', '箱数', '箱规单箱数量', '配送地址',
            '供应商', '货件编号', 'ReferenceId', '箱号', '总箱数',
            '外箱重量(kg)', '外箱总重量(kg)', '外箱长(cm)', '外箱宽(cm)', '外箱高(cm)',
            '外箱体积(m³)', '外箱总体积(m³)', '外箱总体积重(kg)', '创建时间', '发货时间', '物流商'
        ]
        exist_cols = [c for c in keep_cols if c in merged.columns]
        merged = merged[exist_cols]
        merged.rename(columns={
            '箱规单箱数量': '单箱数量',
            '总箱数': '总箱数编号'
        }, inplace=True)
        return merged
    except Exception as e:
        raise Exception(f"数据加载合并失败：{str(e)}")


def get_amazon_warehouse_zone(address):
    """基于映射表精准判断仓库分区"""
    if pd.isna(address) or address == '未知地址' or str(address).strip() == '':
        return '未知分区'
    address_str = str(address).strip().upper()
    code_pattern = re.compile(r'^[A-Z]{3}\d+(-\w+)?')
    code_match = code_pattern.search(address_str)
    if code_match:
        warehouse_code = code_match.group(0).split('-')[0].split('(')[0]
        if warehouse_code in FBA_WAREHOUSE_MAP:
            return FBA_WAREHOUSE_MAP[warehouse_code][1]
    state_match = re.search(r'\b([A-Z]{2})\b', address_str)
    if state_match:
        state = state_match.group(1)
        for code, (code_state, zone) in FBA_WAREHOUSE_MAP.items():
            if code_state == state:
                return zone
    zip_match = re.search(r'\b(\d{5})\b', address_str)
    if zip_match:
        zip_first = zip_match.group(1)[0]
        if zip_first in ['0', '1', '2', '3']:
            return '美东'
        elif zip_first in ['4', '5', '6', '7']:
            return '美中'
        elif zip_first in ['8', '9']:
            return '美西'
    return '未知分区'


def perform_summary(df):
    """按 地址+货件编号+ReferenceId 汇总"""
    df['配送地址'] = df['配送地址'].fillna('未知地址')
    df['货件编号'] = df['货件编号'].fillna('未知编号')
    df['ReferenceId'] = df['ReferenceId'].fillna('无ReferenceId')
    summary = df.groupby(['配送地址', '货件编号', 'ReferenceId'], dropna=False).agg(
        品名=('品名', 'first'),
        供应商=('供应商', 'first'),
        总箱数=('箱数', 'sum'),
        外箱总重量=('外箱总重量(kg)', 'sum'),
        外箱总体积=('外箱总体积(m³)', 'sum')
    ).reset_index()
    summary['仓库分区'] = summary['配送地址'].apply(get_amazon_warehouse_zone)
    col_order = ['仓库分区'] + [col for col in summary.columns if col != '仓库分区']
    summary = summary[col_order]
    summary['外箱总体积'] = summary['外箱总体积'].fillna(0)
    summary['外箱总体积重(kg)'] = (summary['外箱总体积'] * 167).round(2)
    return summary


def save_excel(df, summary_df, save_path):
    """保存为两个 Sheet：详细数据 + 汇总结果"""
    try:
        with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='详细数据', index=False)
            summary_df.to_excel(writer, sheet_name='汇总结果', index=False)
        return save_path
    except Exception as e:
        raise IOError(f"文件保存失败：{str(e)}")


# ===================== 对外主入口 =====================
def process(input_path, output_path, log=None, progress=None):
    """
    FBA 处理主流程
    :param input_path:  上传的发货单 Excel 路径
    :param output_path: 处理结果保存路径
    :param log: 日志回调函数
    :param progress: 进度回调函数 0-100
    :return: output_path
    """
    def emit(msg):
        (log or print)(msg)

    def prog(p, msg=""):
        if progress:
            progress(p, msg)
        if msg:
            emit(msg)

    prog(10, "正在加载并合并数据...")
    merged_df = load_and_merge_data(input_path, log=log)
    prog(40, "正在生成汇总数据...")
    summary_df = perform_summary(merged_df)
    prog(70, "正在保存 Excel 文件...")
    output_path = save_excel(merged_df, summary_df, output_path)
    prog(85, "正在美化表格格式...")
    beautify_excel(output_path, log=log)
    prog(100, f"✅ 处理完成: {output_path}")
    return output_path
